#!/usr/bin/env python3
"""
Import Excel data into Vengage DB.

Reads two sheets from the Excel file:
  - RAW Data-Imaging  → creates Center records linked to matched Customers
  - Invoice Data      → creates CustomerProductAndService records with rates

Usage:
    cd backend

    # Dry-run — shows what will match/fail without touching the DB
    python scripts/import_excel_data.py --validate

    # Actually import
    python scripts/import_excel_data.py

    # Custom file path
    python scripts/import_excel_data.py /path/to/file.xlsx
    python scripts/import_excel_data.py --validate /path/to/file.xlsx

Outputs a summary to stdout and writes unmatched/failed rows to errors_import.csv
in the current working directory.
"""

import csv
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

from app.db.session import SessionLocal
import app.models  # noqa: F401 — registers all models so SQLAlchemy can resolve relationships
from app.models.center import Center
from app.models.customer import Customer
from app.models.customer_product_and_service import CustomerProductAndService
from app.models.product_and_service import ProductAndService

RAW_SHEET = "RAW Data-Imaging"
INVOICE_SHEET = "Invoice Data"
INVOICE_DATA_HEADER_ROW_INDEX = 12  # 0-based; row 13 in Excel is the column header


# ── Matching helpers ───────────────────────────────────────────────────────────

_PAREN_SUFFIX = re.compile(r"\s*\(.*?\)\s*$")


def _match_customer_exact(customers: list, name: str):
    """
    Match by exact display_name, then retry after stripping trailing
    parenthetical codes from the DB name (e.g. 'The Radiology Centre (TRC)'
    matches sheet name 'The Radiology Centre').
    """
    name_l = name.strip().lower()
    fallback = []
    for c in customers:
        dn = c.display_name.strip()
        if dn.lower() == name_l:
            return c
        stripped = _PAREN_SUFFIX.sub("", dn).strip().lower()
        if stripped == name_l:
            fallback.append(c)
    return fallback[0] if fallback else None


# Generic suffix/connector tokens that appear in prefix codes but carry no customer identity.
# Blocking them prevents false matches against customers whose names happen to contain these words.
_GENERIC_PREFIX_TOKENS = frozenset({"ALL", "REG", "DUMMY"})


def _match_customer_by_prefix(customers: list, prefix: str):
    """
    Return the first customer whose display_name contains the prefix as a discrete token.

    Builds two token sets and checks both:
    - Without hyphen splitting: keeps SM-TGC as one token so 'SM-TGC' matches
    - With hyphen splitting:    breaks CDI-Castle into CDI + Castle so 'CDI' matches

    Also checks each hyphen-segment of the prefix itself (e.g. 'SCMI' from 'SCMI-BA')
    so that customer '(SCMI) Southern Cross Medical Imaging' matches prefix 'SCMI-BA'.

    Single-character and generic connector segments (ALL, REG, DUMMY) are excluded from
    split candidates to prevent false matches on unrelated customers.
    """
    prefix_up = prefix.strip().upper()
    split_candidates = {
        p for p in prefix_up.split("-")
        if len(p) > 1 and p not in _GENERIC_PREFIX_TOKENS
    }
    candidates = {prefix_up} | split_candidates

    for c in customers:
        tokens = {t.upper() for t in re.split(r"[\s\(\)\&\,\/\+\.]+", c.display_name) if t.strip()}
        tokens |= {t.upper() for t in re.split(r"[\s\(\)\&\,\/\+\.\-]+", c.display_name) if t.strip()}
        if candidates & tokens:
            return c
    return None


_LEGAL_SUFFIXES = re.compile(
    r"\s*(pty\.?\s*ltd\.?|ltd\.?|pty\.?|inc\.?|llc\.?|co\.?)$",
    re.IGNORECASE,
)


def _strip_legal_suffix(name: str) -> str:
    return _LEGAL_SUFFIXES.sub("", name).strip(" ,.-")


_GENERIC_WORDS = {"medical", "imaging", "radiology", "diagnostics", "diagnostic",
                  "health", "centre", "center", "clinic", "services", "all", "and"}

# Manual overrides: center_code → customer display_name (or None to skip the center entirely).
# These take full priority over automatic prefix/name matching.
_MANUAL_CENTER_OVERRIDES: dict[str, str | None] = {
    "VNG-IMG-9":    "Mi Scan Radiology",
    "VNG-IMG-7-A":  "Carewell Diagnostix",
    "VNG-IMG-7-B":  "Carewell Diagnostix",
    "VNG-IMG-36-A": "DNA Solutions Australia pty ltd (Scanaptics)",
    "VNG-IMG-6-H":  "Synergy Radiology Pty. Ltd.",
    "VNG-IMG-21":   None,
    "VNG-IMG-38":   None,
    "VNG-IMG-37-A": "PINNACLE MEDICAL IMAGING - North Cote",
    "VNG-IMG-54-T": "Vision Radiology",
    "VNG-IMG-42":   "Southwest Radiology",
    "VNG-IMG-44":   "Strategic Care Pty Ltd.",
    "VNG-IMG-36-B": None,
    "VNG-IMG-43-B": "QSCAN Radiology Clinics",
    "VNG-IMG-48":   None,
    "VNG-IMG-53":   "Strategic Care Pty Ltd.",
    "VNG-IMG-56":   "Strategic Care Pty Ltd.",
}


def _find_override_customer(customers: list, name: str):
    """Flexible customer lookup for manual overrides.

    Tries exact match first, then normalizes by stripping legal suffixes and
    trailing parentheticals from both the lookup name and each DB display_name.
    This handles minor formatting differences like 'Pty. Ltd' vs 'Pty. Ltd.'.
    """
    c = _match_customer_exact(customers, name)
    if c:
        return c
    name_norm = _strip_legal_suffix(_PAREN_SUFFIX.sub("", name).strip()).lower()
    for c in customers:
        db_norm = _strip_legal_suffix(_PAREN_SUFFIX.sub("", c.display_name.strip()).strip()).lower()
        if db_norm == name_norm:
            return c
    return None


def _match_customer_by_center_name(customers: list, center_name: str):
    """
    Fallback: find a customer whose display_name matches the center's full name.

    Strategy 1 — customer name in center name:
      Tries full customer name and each segment split by ' - ' / ','
      (e.g. 'Wollongong Diagnostics' from 'ALL CENTERS - Wollongong Diagnostics').
      Requires at least 2 words per segment.
      Winner = customer whose matched segment is longest (most specific match).

    Strategy 2 — brand word from center name in customer name:
      Takes the first significant word of the center name (≥5 chars, not generic)
      and checks if it appears as a token in the customer display_name.
      (e.g. 'CareScan' from 'CareScan Edmondson Park' → 'CareScan Medical Imaging').
    """
    center_name_l = center_name.strip().lower()

    # Strategy 1: track (customer, matched_segment_length) to prefer the most specific hit.
    # Example: "Synergy Radiology" (16 chars) beats "Rouse Hill" (9 chars), so Synergy wins
    # over a customer that only matches via a short geographic suffix.
    seg_matches: list[tuple] = []
    for c in customers:
        dn_full = _strip_legal_suffix(c.display_name.strip())
        segments = [dn_full] + [
            s.strip() for s in re.split(r"\s*[-,]\s*", dn_full) if s.strip()
        ]
        for seg in segments:
            if len(seg.split()) < 2:
                continue
            if seg.lower() in center_name_l:
                seg_matches.append((c, len(seg)))
                break

    if seg_matches:
        return max(seg_matches, key=lambda x: (x[1], len(x[0].display_name)))[0]

    # Strategy 2: first significant word of center name in customer display_name
    matches = []
    center_words = re.split(r"[\s\(\)\&\,\/\+\.\-]+", center_name)
    brand_word = next(
        (w for w in center_words if len(w) >= 5 and w.lower() not in _GENERIC_WORDS),
        None,
    )
    if brand_word:
        brand_up = brand_word.upper()
        for c in customers:
            tokens = {t.upper() for t in re.split(r"[\s\(\)\&\,\/\+\.\-]+", c.display_name) if t.strip()}
            if brand_up in tokens:
                matches.append(c)

    if not matches:
        return None
    return max(matches, key=lambda c: len(c.display_name))


def _match_product_exact(products: list, name: str):
    name_l = name.strip().lower()
    for p in products:
        if p.name.strip().lower() == name_l:
            return p
    return None


# ── Sheet parsers ──────────────────────────────────────────────────────────────

def _parse_raw_sheet(ws) -> list[tuple[str, str, list[str]]]:
    """Returns list of (center_code, center_name, [prefix1, prefix2, ...])."""
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        center_code = str(row[0]).strip()
        center_name = str(row[1]).strip() if row[1] else ""
        raw_prefix = str(row[2]).strip() if row[2] else ""
        prefixes = [p.strip() for p in raw_prefix.split(",") if p.strip()]
        if center_code:
            result.append((center_code, center_name, prefixes))
    return result


def _parse_invoice_sheet(ws) -> list[tuple[str, str, float]]:
    """
    Returns unique list of (customer_name, product_name, rate).
    Handles continuation rows (no customer name = same customer as previous row).
    """
    rows = list(ws.iter_rows(values_only=True))
    result = []
    seen: set[tuple[str, str]] = set()
    current_customer: str | None = None

    for row in rows[INVOICE_DATA_HEADER_ROW_INDEX + 1:]:
        if row[1]:
            current_customer = str(row[1]).strip()
        if not current_customer:
            continue
        product = row[7]
        rate = row[10]
        if not product or rate is None:
            continue
        product_name = str(product).strip()
        try:
            rate_f = float(rate)
        except (TypeError, ValueError):
            continue
        key = (current_customer, product_name)
        if key not in seen:
            seen.add(key)
            result.append((current_customer, product_name, rate_f))

    return result


# ── Validate (dry-run) ─────────────────────────────────────────────────────────

def _run_validate(excel_path: Path) -> None:
    print(f"Reading: {excel_path.name}")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    raw_centers = _parse_raw_sheet(wb[RAW_SHEET])
    invoice_combos = _parse_invoice_sheet(wb[INVOICE_SHEET])

    print(f"  {len(raw_centers)} centers in RAW sheet")
    print(f"  {len(invoice_combos)} unique customer-product rows in Invoice Data sheet")

    db = SessionLocal()
    try:
        db_customers = db.query(Customer).all()
        db_products = (
            db.query(ProductAndService)
            .filter(ProductAndService.active == True)
            .all()
        )
        existing_centers = {c.name for c in db.query(Center).all()}
        existing_cps = {
            (r.customer_id, r.product_and_service_id)
            for r in db.query(CustomerProductAndService).all()
        }

        print(f"\nDB: {len(db_customers)} customers, {len(db_products)} active products, "
              f"{len(existing_centers)} existing centers\n")

        # ── Centers preview ────────────────────────────────────────────────────
        matched_centers: list[tuple[str, str]] = []
        unmatched_centers: list[tuple[str, str]] = []
        already_exist_centers: list[str] = []
        skipped_centers: list[str] = []

        for center_code, center_name, prefixes in raw_centers:
            if center_code in existing_centers:
                already_exist_centers.append(center_code)
                continue
            if center_code in _MANUAL_CENTER_OVERRIDES:
                override_name = _MANUAL_CENTER_OVERRIDES[center_code]
                if override_name is None:
                    skipped_centers.append(center_code)
                    continue
                matched_customer = _find_override_customer(db_customers, override_name)
            else:
                matched_customer = None
                for prefix in prefixes:
                    matched_customer = _match_customer_by_prefix(db_customers, prefix)
                    if matched_customer:
                        break
                if not matched_customer and center_name:
                    matched_customer = _match_customer_by_center_name(db_customers, center_name)
            if matched_customer:
                matched_centers.append((center_code, matched_customer.display_name))
            else:
                unmatched_centers.append((center_code, ", ".join(prefixes)))

        print("── Centers (RAW Data-Imaging) ──────────────────────────────────────")
        print(f"  Will be created : {len(matched_centers)}")
        print(f"  Already in DB   : {len(already_exist_centers)}")
        print(f"  Skipped (manual): {len(skipped_centers)}")
        print(f"  No match found  : {len(unmatched_centers)}")

        if already_exist_centers:
            print("\n  ALREADY IN DB:")
            for code in already_exist_centers:
                print(f"    {code}")

        if skipped_centers:
            print("\n  SKIPPED (manual override — no DB record intended):")
            for code in skipped_centers:
                print(f"    {code}")

        if matched_centers:
            print("\n  WILL CREATE:")
            for code, cust in matched_centers:
                print(f"    {code:22s} → {cust}")

        if unmatched_centers:
            print("\n  NO MATCH (customer not in DB):")
            for code, prefixes in unmatched_centers:
                print(f"    {code:22s}   prefixes: {prefixes}")

        # ── Customer-Product-Service preview ───────────────────────────────────
        cps_will_create: list[tuple[str, str, float]] = []
        cps_already_exist: list[tuple[str, str]] = []
        cps_no_customer: list[tuple[str, str]] = []
        cps_no_product: list[tuple[str, str, str]] = []

        for customer_name, product_name, rate in invoice_combos:
            customer = _match_customer_exact(db_customers, customer_name)
            if not customer:
                cps_no_customer.append((customer_name, product_name))
                continue
            product = _match_product_exact(db_products, product_name)
            if not product:
                cps_no_product.append((customer_name, product_name, customer.display_name))
                continue
            if (customer.id, product.id) in existing_cps:
                cps_already_exist.append((customer_name, product_name))
            else:
                cps_will_create.append((customer_name, product_name, rate))

        print(f"\n── Customer → Product/Service rates (Invoice Data) ─────────────────")
        print(f"  Will be created : {len(cps_will_create)}")
        print(f"  Already in DB   : {len(cps_already_exist)}")
        print(f"  Customer not found : {len(set(c for c, _ in cps_no_customer))}"
              f" customers ({len(cps_no_customer)} rows)")
        print(f"  Product not found  : {len(set(p for _, p, _ in cps_no_product))} products"
              f" ({len(cps_no_product)} rows)")

        if cps_will_create:
            print("\n  WILL CREATE:")
            prev_cust = None
            for cust, prod, rate in cps_will_create:
                if cust != prev_cust:
                    print(f"    [{cust}]")
                    prev_cust = cust
                print(f"      {prod}  @ {rate}")

        if cps_no_product:
            unique_missing = sorted({p for _, p, _ in cps_no_product})
            print("\n  PRODUCTS NOT IN DB (sync from QBO first):")
            for p in unique_missing:
                print(f"    {p}")

        if cps_no_customer:
            unique_missing_cust = sorted({c for c, _ in cps_no_customer})
            print("\n  CUSTOMERS NOT IN DB:")
            for c in unique_missing_cust:
                print(f"    {c}")

        total_unmatched = len(unmatched_centers) + len(cps_no_customer) + len(cps_no_product)
        print(f"\n{'─'*60}")
        print(f"Total rows that will be CREATED : "
              f"{len(matched_centers)} centers + {len(cps_will_create)} service rates")
        print(f"Total rows with NO MATCH        : {total_unmatched}")
        print(f"\nRun without --validate to apply.")

    finally:
        db.close()


# ── Import ─────────────────────────────────────────────────────────────────────

def _run_import(excel_path: Path) -> None:
    print(f"Reading: {excel_path.name}")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    raw_centers = _parse_raw_sheet(wb[RAW_SHEET])
    invoice_combos = _parse_invoice_sheet(wb[INVOICE_SHEET])

    print(f"  {len(raw_centers)} centers in RAW sheet")
    print(f"  {len(invoice_combos)} unique customer-product rows in Invoice Data sheet")

    db = SessionLocal()
    errors: list[dict] = []

    try:
        db_customers = db.query(Customer).all()
        db_products = (
            db.query(ProductAndService)
            .filter(ProductAndService.active == True)
            .all()
        )
        print(f"\nDB: {len(db_customers)} customers, {len(db_products)} active products\n")

        # ── Step 1: CustomerProductAndService ──────────────────────────────────
        print("── Step 1: Customer → Product/Service rates ──")
        cps_created = cps_skipped = 0

        for customer_name, product_name, rate in invoice_combos:
            customer = _match_customer_exact(db_customers, customer_name)
            if not customer:
                errors.append({
                    "sheet": INVOICE_SHEET,
                    "type": "customer_not_found",
                    "customer": customer_name,
                    "product": product_name,
                    "rate": rate,
                    "reason": "No DB customer matched this name",
                })
                continue

            product = _match_product_exact(db_products, product_name)
            if not product:
                errors.append({
                    "sheet": INVOICE_SHEET,
                    "type": "product_not_found",
                    "customer": customer_name,
                    "product": product_name,
                    "rate": rate,
                    "reason": "No active product/service in DB matched this name",
                })
                continue

            existing = (
                db.query(CustomerProductAndService)
                .filter_by(customer_id=customer.id, product_and_service_id=product.id)
                .first()
            )
            if existing:
                cps_skipped += 1
                continue

            db.add(CustomerProductAndService(
                customer_id=customer.id,
                product_and_service_id=product.id,
                rate=rate,
            ))
            cps_created += 1

        db.flush()
        print(f"  Created: {cps_created}  |  Already existed: {cps_skipped}")

        # ── Step 2: Centers ────────────────────────────────────────────────────
        print("\n── Step 2: Center codes → Customers ──")
        centers_created = centers_skipped = 0

        for center_code, center_name, prefixes in raw_centers:
            if center_code in _MANUAL_CENTER_OVERRIDES:
                override_name = _MANUAL_CENTER_OVERRIDES[center_code]
                if override_name is None:
                    print(f"  {center_code:20s}   [skipped — manual override]")
                    continue
                matched_customer = _find_override_customer(db_customers, override_name)
                if not matched_customer:
                    errors.append({
                        "sheet": RAW_SHEET,
                        "type": "center_override_customer_not_found",
                        "customer": override_name,
                        "product": "",
                        "rate": "",
                        "reason": f"Center '{center_code}' — manual override customer '{override_name}' not found in DB",
                    })
                    continue
            else:
                matched_customer = None
                for prefix in prefixes:
                    matched_customer = _match_customer_by_prefix(db_customers, prefix)
                    if matched_customer:
                        break
                if not matched_customer and center_name:
                    matched_customer = _match_customer_by_center_name(db_customers, center_name)

            if not matched_customer:
                errors.append({
                    "sheet": RAW_SHEET,
                    "type": "center_no_customer_match",
                    "customer": "",
                    "product": "",
                    "rate": "",
                    "reason": f"Center '{center_code}' — no customer matched prefixes: {', '.join(prefixes)}",
                })
                continue

            existing = db.query(Center).filter_by(name=center_code).first()
            if existing:
                centers_skipped += 1
                continue

            db.add(Center(name=center_code, company_id=matched_customer.id))
            centers_created += 1
            print(f"  {center_code:20s} → {matched_customer.display_name}")

        db.commit()
        print(f"\n  Created: {centers_created}  |  Already existed: {centers_skipped}")

        # ── Step 3: Error CSV ──────────────────────────────────────────────────
        errors_path = Path.cwd() / "errors_import.csv"
        if errors:
            with open(errors_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=["sheet", "type", "customer", "product", "rate", "reason"],
                )
                writer.writeheader()
                writer.writerows(errors)
            print(f"\nWARNING: {len(errors)} unmatched rows → {errors_path}")
        else:
            print("\nNo errors.")

        print("\nImport complete.")

    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


# ── Entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    args = sys.argv[1:]
    validate_mode = "--validate" in args
    file_args = [a for a in args if a != "--validate"]

    default_path = (
        Path(__file__).parent.parent.parent
        / "Intuit-Invoice-Data-Generation-STEP-1-MARCH-26-backup (1).xlsx"
    )
    path = Path(file_args[0]) if file_args else default_path

    if validate_mode:
        _run_validate(path)
    else:
        _run_import(path)

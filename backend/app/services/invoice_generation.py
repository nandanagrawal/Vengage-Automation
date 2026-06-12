"""Generate QBO invoices from an uploaded RAW Data-Imaging XLSX/XLS/CSV file.

New file format (RAW Data-Imaging sheet)
-----------------------------------------
Row 0   : header row — column names like "Confirmed Appointment (4)", etc.
Col 0   : S.No. / Center ID  — matched case-insensitively against Center.name
          in the database (e.g. "VNG-IMG-1-A").
Col 1   : Center Name        — human-readable label, ignored for processing.
Col 2   : Center Prefix      — may be comma-separated ("PAR, SMI-ALL"); the
          first token before any comma is used in ItemDescription.
Col 3+  : metric quantity columns

Center matching
---------------
The value in col 0 is matched case-insensitively against Center.name in the
database.  Rows whose col-0 value does not match any Center are skipped
(recorded in errors).

Product → column mapping
------------------------
PRODUCT_COLUMN_MAP maps each ProductAndService name (lower-case) to one or
more column-header substrings.  The line-item quantity is the sum of all
matching column values for that center's row.  Products whose name is in
FIXED_QUANTITY_PRODUCTS always receive quantity = 1.

Invoice grouping
----------------
Existing Invoice groupings define which centers produce one combined invoice.
Centers NOT in any grouping get one individual invoice each.

Critically: grouped centers each produce their OWN set of line items inside
the combined invoice (quantities are NOT summed across centers).

Invoice fields
--------------
  TxnDate / ServiceDate : last calendar day of the current month
  DueDate               : TxnDate + 15 days
  Memo (CustomerMemo)   : "{MMMYY} Invoice"  e.g. "MAR26 Invoice"
  Rate                  : CustomerProductAndService.rate  (per-customer, per-product)
  ItemDescription       : "{center_prefix} - {MMMYY} Invoice month {service_code}"
  0-quantity items      : included (amount = 0)
"""

from __future__ import annotations

import calendar
import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session, selectinload

from app.core.config import settings
from app.models.center import Center
from app.models.customer import Customer
from app.models.customer_product_and_service import CustomerProductAndService
from app.models.generated_invoice import (
    GeneratedInvoice,
    GeneratedInvoiceCenter,
    GeneratedInvoiceLineItem,
)
from app.models.invoice import Invoice
from app.models.product_and_service import ProductAndService
from app.services.qbo_client import SupportsQuickBooks

# Module-level cache: maps "realm_id:code_name" → resolved QBO TaxCode Id
_tax_code_id_cache: dict[str, str] = {}


def _resolve_tax_code_id(qbo: SupportsQuickBooks, token: str, realm: str, code: str) -> str:
    """Resolve a tax code name (e.g. 'GST') to its QBO Id (e.g. '5').

    If `code` is already numeric it is returned as-is.
    Falls back to the original value if the lookup fails.
    """
    if code.isdigit():
        return code
    cache_key = f"{realm}:{code.lower()}"
    if cache_key in _tax_code_id_cache:
        return _tax_code_id_cache[cache_key]
    try:
        codes = qbo.query_tax_codes(token, realm)
        for tc in codes:
            if str(tc.get("Name", "")).lower() == code.lower():
                resolved = str(tc["Id"])
                _tax_code_id_cache[cache_key] = resolved
                return resolved
    except Exception:
        pass
    return code


# ── Product → metric-column mapping ──────────────────────────────────────────
#
# Keys   : ProductAndService.name lowercased (partial or full match via `in`)
# Values : list of column-header substrings (lowercased); the quantity is the
#          sum of all columns whose lowercased header contains one of these substrings.
#
# Multi-column entries (e.g. Call Forwarding) are summed together.

PRODUCT_COLUMN_MAP: dict[str, list[str]] = {
    "olivia ai bookings for imaging workflow": ["confirmed appointment (4)"],
    "olivia ai - provisional bookings for imaging workflow": ["provisional appointment (5)"],
    "direct calls handling charges": ["direct call (6)"],
    "e-referral transmission charges": ["total e-referrals (11)"],
    "e-referral manual sms charges": ["e-referral manual sms count (13)"],
    "bookings via specialist portal": ["appointments (specialist) (15)"],
    "e-referral greeting sms charges": ["e-referral greeting sms count (12)"],
    "bookings directly done by customer team": ["appointments (operators) (14)"],
    "misc item": ["sms count (chat) (16)"],
    "e-referral portal monthly subscription fees": ["total e-referrals (11)"],
    "call forwarding - telephony charges": [
        "voice call forwarding bh (mins) (9)",
        "voice call forwarding ooh (mins) (10)",
    ],
    "olivia ai - walkin services call handling": ["walkin (7)"],
    "internal e-referral charges": ["total e-referrals (11)"],
    "external e-referral charges": ["e-referral greeting sms count (12)"],
}

# Products that always have quantity = 1 regardless of file data
FIXED_QUANTITY_PRODUCTS: frozenset[str] = frozenset({
    "kiosk monthly usage and support charges",
})


# ── Parsing ───────────────────────────────────────────────────────────────────

@dataclass
class ParsedFile:
    """center_name_lower (col 0) → {col_header_lower: Decimal quantity}"""
    rows: dict[str, dict[str, Decimal]] = field(default_factory=dict)
    # original-case column headers (col 3+ only)
    metric_columns: list[str] = field(default_factory=list)
    # center_name_lower → first token of col 2 (used in line-item descriptions)
    center_prefixes: dict[str, str] = field(default_factory=dict)
    # center_name_lower → original-case col 0 value
    center_display_names: dict[str, str] = field(default_factory=dict)
    # center_name_lower → col 1 value (Center Name column)
    center_col1_names: dict[str, str] = field(default_factory=dict)


def _to_decimal(value: Any) -> Decimal:
    if value is None or (isinstance(value, str) and not value.strip()):
        return Decimal("0")
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        return Decimal("0")


def _extract_center_prefix(raw: Any) -> str:
    """Return the first comma-token of a Center Prefix cell, stripped."""
    text = str(raw).strip() if raw is not None else ""
    return text.split(",")[0].strip()


def _parse_raw_data_rows(
    headers: list[str],
    data_rows: list[list[Any]],
) -> ParsedFile:
    """Parse RAW Data-Imaging sheet.

    Expects:
      col 0 = Center ID / S.No.  → matched against Center.name in DB
      col 1 = Center Name        → ignored
      col 2 = Center Prefix      → first comma-token used in descriptions
      col 3+ = metric columns
    """
    if len(headers) < 3:
        raise ValueError("File must have at least 3 columns (S.No., Center Name, Center Prefix).")

    metric_headers = headers[3:]  # col 3 onwards
    result = ParsedFile(metric_columns=metric_headers)

    for row in data_rows:
        if not row or len(row) < 1:
            continue
        center_name = str(row[0]).strip() if row[0] is not None else ""
        if not center_name:
            continue

        center_name_lower = center_name.lower()
        # col 2 first token is used for line-item descriptions (fallback: center_name)
        desc_prefix = _extract_center_prefix(row[2]) if len(row) > 2 else ""
        if not desc_prefix:
            desc_prefix = center_name

        metrics: dict[str, Decimal] = {}
        for i, col_header in enumerate(metric_headers, start=3):
            raw_val = row[i] if i < len(row) else None
            metrics[col_header.lower()] = _to_decimal(raw_val)

        col1_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        # Accumulate if the same center appears on multiple rows
        if center_name_lower in result.rows:
            existing = result.rows[center_name_lower]
            for k, v in metrics.items():
                existing[k] = existing.get(k, Decimal("0")) + v
        else:
            result.rows[center_name_lower] = metrics
            result.center_prefixes[center_name_lower] = desc_prefix
            result.center_display_names[center_name_lower] = center_name
            result.center_col1_names[center_name_lower] = col1_name

    return result


def _load_xlsx_sheet(content: bytes, sheet_name: str = "RAW Data-Imaging") -> tuple[list[str], list[list[Any]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for .xlsx files.") from exc

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    # Prefer the named sheet; fall back to active
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        raise ValueError("XLSX file is empty.")
    headers = [str(h).strip() if h is not None else "" for h in rows_raw[0]]
    return headers, [list(r) for r in rows_raw[1:]]


def parse_xlsx(content: bytes) -> ParsedFile:
    headers, data_rows = _load_xlsx_sheet(content)
    return _parse_raw_data_rows(headers, data_rows)


def parse_xls(content: bytes) -> ParsedFile:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required for .xls files.") from exc

    wb = xlrd.open_workbook(file_contents=content)
    # Prefer "RAW Data-Imaging" sheet
    sheet_names = [wb.sheet_by_index(i).name for i in range(wb.nsheets)]
    idx = sheet_names.index("RAW Data-Imaging") if "RAW Data-Imaging" in sheet_names else 0
    ws = wb.sheet_by_index(idx)
    if ws.nrows == 0:
        raise ValueError("XLS file is empty.")
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    data_rows = [
        [ws.cell_value(r, c) for c in range(ws.ncols)]
        for r in range(1, ws.nrows)
    ]
    return _parse_raw_data_rows(headers, data_rows)


def parse_csv(content: bytes) -> ParsedFile:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV file is empty.")
    headers = [str(h).strip() for h in rows[0]]
    return _parse_raw_data_rows(headers, rows[1:])


def parse_spreadsheet(filename: str, content: bytes) -> ParsedFile:
    name_lower = filename.lower()
    if name_lower.endswith(".csv"):
        return parse_csv(content)
    elif name_lower.endswith(".xlsx"):
        return parse_xlsx(content)
    elif name_lower.endswith(".xls"):
        return parse_xls(content)
    raise ValueError(
        f"Unsupported file type '{filename}'. Upload a .csv, .xlsx, or .xls file."
    )


# ── Date helpers ──────────────────────────────────────────────────────────────

def _invoice_date() -> date:
    today = date.today()
    first_of_current = today.replace(day=1)
    last_month_last_day = first_of_current - timedelta(days=1)
    return last_month_last_day


def _due_date(inv_date: date) -> date:
    return inv_date + timedelta(days=15)


def _memo(inv_date: date) -> str:
    month = inv_date.strftime("%B")
    year = inv_date.strftime("%Y")
    return (
        f"Thank you for your business and for being our valuable customer. "
        f"Please find the details for the {month}, {year} month bookings.\n\n"
        f"Hope you have a great day ahead."
    )


def _month_label(inv_date: date) -> str:
    """e.g. MAR26"""
    return inv_date.strftime("%b%y").upper()


# ── Slab pricing helpers ──────────────────────────────────────────────────────

# Matches "(Slab-1)", "(Slab-2b)", "(slab_3a)" etc. anywhere in the name.
_SLAB_SUFFIX_RE = re.compile(r"\s*\(\s*slab[-_]?\s*\d+[a-z]*\s*\)", re.IGNORECASE)
# Matches trailing service-code tokens like "B0001", "B0002", "B0003".
_SERVICE_CODE_RE = re.compile(r"\s+b\d{4}\s*$", re.IGNORECASE)
# Bounded range: "1-1000", "1001-2500" (en-dash or hyphen).
_RANGE_BOUNDED_RE = re.compile(r"(\d[\d,]*)\s*[-–]\s*(\d[\d,]*)")
# Unbounded range: "2501+" or "2501 +".
_RANGE_UNBOUNDED_RE = re.compile(r"(\d[\d,]*)\s*\+")


def _get_col_substrings(product_name_lower: str) -> list[str] | None:
    """Return PRODUCT_COLUMN_MAP column-substrings for a product name.

    Tries four normalizations in order so that products whose names include
    a slab suffix ("(Slab-1b)") and/or a service-code suffix ("B0001") still
    resolve to the correct column mapping:

      1. Exact name (no transformation)
      2. Strip service-code suffix
      3. Strip slab suffix
      4. Strip both slab and service-code suffixes
    """
    if product_name_lower in PRODUCT_COLUMN_MAP:
        return PRODUCT_COLUMN_MAP[product_name_lower]

    no_code = _SERVICE_CODE_RE.sub("", product_name_lower).strip()
    if no_code != product_name_lower and no_code in PRODUCT_COLUMN_MAP:
        return PRODUCT_COLUMN_MAP[no_code]

    no_slab = _SLAB_SUFFIX_RE.sub("", product_name_lower).strip()
    if no_slab != product_name_lower and no_slab in PRODUCT_COLUMN_MAP:
        return PRODUCT_COLUMN_MAP[no_slab]

    no_both = _SERVICE_CODE_RE.sub("", no_slab).strip()
    if no_both != product_name_lower and no_both in PRODUCT_COLUMN_MAP:
        return PRODUCT_COLUMN_MAP[no_both]

    return None


def _parse_slab_range(description: str | None) -> tuple[int, int | None] | None:
    """Parse a numeric range from a ProductAndService description.

    Returns ``(start, end)`` for bounded ranges (e.g. "1-1000" → (1, 1000))
    or ``(start, None)`` for unbounded ranges (e.g. "2501+" → (2501, None)).
    Returns ``None`` when no recognisable range is found.

    Handles:
      - "1-1000 bookings slab"
      - "1001–2500"   (en-dash)
      - "2501+ bookings"
      - "2,501+"      (comma-grouped numbers)
    """
    if not description:
        return None
    m = _RANGE_BOUNDED_RE.search(description)
    if m:
        start = int(m.group(1).replace(",", ""))
        end = int(m.group(2).replace(",", ""))
        return (start, end)
    m = _RANGE_UNBOUNDED_RE.search(description)
    if m:
        start = int(m.group(1).replace(",", ""))
        return (start, None)
    return None


def _slab_qty(total: Decimal, start: int, end: int | None) -> Decimal:
    """Compute appointments falling within slab [start, end].

    The slab covers appointment numbers start through end (inclusive).
    ``end=None`` means unbounded (covers all appointments from start onward).

    Examples (total=2755):
      slab (1, 1000)   → 1000
      slab (1001, 2500) → 1500
      slab (2501, None) → 255

    Zero is returned when total < start.
    """
    t = int(total)
    below = start - 1          # appointments before this slab
    if t <= below:
        return Decimal("0")
    if end is None:
        return Decimal(str(t - below))
    return Decimal(str(min(t, end) - below))


# ── Quantity lookup ───────────────────────────────────────────────────────────

def _get_quantity(
    product_name_lower: str,
    center_metrics: dict[str, Decimal],
) -> Decimal | None:
    """Look up quantity for a product from a center's metric row.

    Returns None if there is no mapping at all (product should be skipped).
    Returns Decimal (possibly 0) when a mapping exists but has zero value.
    """
    if product_name_lower in FIXED_QUANTITY_PRODUCTS:
        return Decimal("1")

    col_substrings = _get_col_substrings(product_name_lower)
    if col_substrings:
        total = Decimal("0")
        for substr in col_substrings:
            for col_lower, val in center_metrics.items():
                if substr in col_lower:
                    total += val
        return total

    # Fallback: column header contains the product name as a substring.
    # Allows custom/future products and test data to work without an explicit entry.
    for col_lower, val in center_metrics.items():
        if product_name_lower in col_lower:
            return val

    return None  # no mapping → line item skipped


# ── Invoice generation ────────────────────────────────────────────────────────

@dataclass
class _LineItem:
    product_and_service_id: int | None
    product_name: str
    center_prefix: str
    description: str
    quantity: Decimal
    rate: Decimal
    amount: Decimal
    qbo_payload: dict[str, Any]


@dataclass
class InvoiceDetail:
    customer_display_name: str
    group_description: str
    qbo_invoice_id: str
    invoice_number: str | None
    sent_at: str | None
    send_status: str
    generated_invoice_id: int | None = None

    @property
    def sent(self) -> bool:
        return self.send_status == "sent"


@dataclass
class GenerationResult:
    total_center_rows: int = 0
    centers_matched: int = 0
    centers_skipped: int = 0
    invoices_created: int = 0
    invoices_failed: int = 0
    invoice_details: list[InvoiceDetail] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "total_center_rows": self.total_center_rows,
            "centers_matched": self.centers_matched,
            "centers_skipped": self.centers_skipped,
            "invoices_created": self.invoices_created,
            "invoices_failed": self.invoices_failed,
            "invoice_details": [
                {
                    "customer": d.customer_display_name,
                    "group": d.group_description,
                    "qbo_invoice_id": d.qbo_invoice_id,
                    "invoice_number": d.invoice_number,
                    "sent_at": d.sent_at,
                    "send_status": d.send_status,
                    "sent": d.sent,
                    "generated_invoice_id": d.generated_invoice_id,
                }
                for d in self.invoice_details
            ],
            "errors": self.errors,
        }


def _build_line_items_for_center(
    center_prefix: str,
    center_metrics: dict[str, Decimal],
    customer_services: list[CustomerProductAndService],
    month_label: str,
    tax_code_id: str = "",
) -> list[_LineItem]:
    """Build _LineItem objects for one center, supporting slab-based pricing.

    Slab logic
    ----------
    When a customer has multiple services that all map to the **same metric
    column** (e.g. "Confirmed Appointment") AND every one of those services
    has a parseable numeric range in its ProductAndService.description (e.g.
    "1-1000", "1001-2500", "2501+"), the total column value is *distributed*
    across the slabs rather than duplicated:

      total=2755, slabs (1-1000), (1001-2500), (2501+)
        → slab-1 qty=1000, slab-2 qty=1500, slab-3 qty=255

    If only some services have parseable ranges, or the group has just one
    service, each service independently receives the full column value.

    Zero-quantity line items are always included (amount=0).
    """
    desc_base = f"{center_prefix} - {month_label} Invoice month".strip()

    def _make_item(cs: CustomerProductAndService, qty: Decimal) -> _LineItem:
        ps = cs.product_and_service
        rate = cs.rate
        amount = (qty * rate).quantize(Decimal("0.01"))
        qbo_payload: dict[str, Any] = {
            "DetailType": "SalesItemLineDetail",
            "Description": desc_base,
            "Amount": float(amount),
            "SalesItemLineDetail": {
                "ItemRef": {"value": ps.qbo_id},
                "Qty": float(qty),
                "UnitPrice": float(rate),
                "TaxCodeRef": {"value": tax_code_id or settings.QBO_LINE_TAX_CODE},
            },
        }
        return _LineItem(
            product_and_service_id=ps.id,
            product_name=ps.name,
            center_prefix=center_prefix,
            description=desc_base,
            quantity=qty,
            rate=rate,
            amount=amount,
            qbo_payload=qbo_payload,
        )

    # ── 1. Group services by which metric column they read from ───────────────
    # key → sorted join of column substrings (or special prefix for fixed/fallback)
    col_groups: dict[str, list[CustomerProductAndService]] = {}

    for cs in customer_services:
        ps = cs.product_and_service
        if cs.rate is None or cs.rate <= 0:
            continue
        name_lower = ps.name.lower()

        if name_lower in FIXED_QUANTITY_PRODUCTS:
            col_groups.setdefault("__fixed__", []).append(cs)
            continue

        col_subs = _get_col_substrings(name_lower)
        if col_subs is not None:
            key = "|".join(sorted(col_subs))
            col_groups.setdefault(key, []).append(cs)
        else:
            # Fallback: col header contains the product name
            for col_lower in center_metrics:
                if name_lower in col_lower:
                    col_groups.setdefault(f"__fb__{name_lower}", []).append(cs)
                    break
            # If still no match, product has no column mapping → skipped

    # ── 2. Build line items per group ─────────────────────────────────────────
    items: list[_LineItem] = []

    for key, group in col_groups.items():
        # Resolve raw total for this column group
        if key == "__fixed__":
            raw_total = Decimal("1")
        elif key.startswith("__fb__"):
            pname = key[len("__fb__"):]
            raw_total = sum(
                (v for col, v in center_metrics.items() if pname in col),
                Decimal("0"),
            )
        else:
            col_subs = key.split("|")
            raw_total = Decimal("0")
            for substr in col_subs:
                for col_lower, val in center_metrics.items():
                    if substr in col_lower:
                        raw_total += val

        # Check for slab group: ≥2 services that ALL have parseable description ranges
        if len(group) > 1:
            slab_pairs = [
                (cs, _parse_slab_range(cs.product_and_service.description))
                for cs in group
            ]
            if all(r is not None for _, r in slab_pairs):
                # Full slab mode: sort by range start and distribute total
                slab_pairs.sort(key=lambda x: x[1][0])  # type: ignore[index]
                for cs, (start, end) in slab_pairs:  # type: ignore[misc]
                    qty = _slab_qty(raw_total, start, end)
                    items.append(_make_item(cs, qty))
                continue
            # Partial or no parseable ranges → fall through to full-qty mode

        # Non-slab (or partial ranges): each service gets the full raw total
        for cs in group:
            items.append(_make_item(cs, raw_total))

    return items


def _build_qbo_invoice_payload(
    customer: Customer,
    center_names: list[str],
    parsed: ParsedFile,
    center_by_name: dict[str, Center],
    customer_services: list[CustomerProductAndService],
    inv_date: date,
    tax_code_id: str = "",
) -> tuple[dict[str, Any], Decimal, list[_LineItem]] | None:
    """Build the QBO invoice payload with per-center line items."""
    month_label = _month_label(inv_date)
    due = _due_date(inv_date)
    memo = _memo(inv_date)

    all_line_items: list[_LineItem] = []

    for name in center_names:
        ctr = center_by_name.get(name.lower())
        if ctr is None:
            continue
        name_lower = name.lower()
        # Metrics keyed by col-0 value (= Center.name lower)
        center_metrics = parsed.rows.get(name_lower, {})
        # Description prefix = first token of col-2; fall back to center name
        center_prefix = parsed.center_prefixes.get(name_lower, ctr.name)

        items = _build_line_items_for_center(
            center_prefix=center_prefix,
            center_metrics=center_metrics,
            customer_services=customer_services,
            month_label=month_label,
            tax_code_id=tax_code_id,
        )
        all_line_items.extend(items)

    if not all_line_items:
        return None

    total = sum(li.amount for li in all_line_items)

    payload: dict[str, Any] = {
        "CustomerRef": {"value": customer.qbo_id},
        "TxnDate": inv_date.isoformat(),
        "DueDate": due.isoformat(),
        "CustomerMemo": {"value": memo},
        "GlobalTaxCalculation": "TaxExcluded",
        "Line": [li.qbo_payload for li in all_line_items],
    }
    if customer.primary_email:
        payload["BillEmail"] = {"Address": customer.primary_email}
    return payload, total, all_line_items


def generate_invoices_from_parsed(
    db: Session,
    qbo: SupportsQuickBooks,
    access_token: str,
    realm_id: str,
    parsed: ParsedFile,
    invoice_upload_id: int | None = None,
) -> GenerationResult:
    """Run invoice generation from a pre-built ParsedFile (skips file parsing)."""
    result = GenerationResult()

    # Resolve the configured tax code name (e.g. "GST") to its QBO Id (e.g. "5").
    # Australian QBO requires the numeric Id, not the name string.
    tax_code_id = _resolve_tax_code_id(qbo, access_token, realm_id, settings.QBO_LINE_TAX_CODE)

    if not parsed.rows:
        result.errors.append("File contains no data rows.")
        return result

    result.total_center_rows = len(parsed.rows)

    names_in_file = list(parsed.rows.keys())
    centers_in_db: list[Center] = (
        db.query(Center)
        .filter(sa_func.lower(Center.name).in_(names_in_file))
        .all()
    )
    center_by_name: dict[str, Center] = {c.name.lower(): c for c in centers_in_db}

    matched_names: list[str] = []
    for name_lower in names_in_file:
        if name_lower in center_by_name:
            matched_names.append(name_lower)
            result.centers_matched += 1
        else:
            result.centers_skipped += 1
            result.errors.append(f"Center '{name_lower}' not found in database — row skipped.")

    if not matched_names:
        result.errors.append("No centers in the file matched any center in the database.")
        return result

    customer_ids: set[int] = {center_by_name[n].company_id for n in matched_names}
    customers: list[Customer] = (
        db.query(Customer)
        .filter(Customer.id.in_(customer_ids))
        .options(
            selectinload(Customer.customer_services)
            .selectinload(CustomerProductAndService.product_and_service),
        )
        .all()
    )
    customer_by_id: dict[int, Customer] = {c.id: c for c in customers}

    invoices: list[Invoice] = (
        db.query(Invoice)
        .options(selectinload(Invoice.centers))
        .filter(Invoice.company_id.in_(customer_ids))
        .all()
    )
    invoices_by_company: dict[int, list[Invoice]] = {}
    for inv in invoices:
        invoices_by_company.setdefault(inv.company_id, []).append(inv)

    inv_date = _invoice_date()

    for company_id in customer_ids:
        customer = customer_by_id.get(company_id)
        if not customer:
            continue
        if not customer.qbo_id:
            result.errors.append(
                f"Customer '{customer.display_name}' has no QBO ID — skipped (sync first)."
            )
            continue

        customer_centers = [n for n in matched_names if center_by_name[n].company_id == company_id]
        if not customer_centers:
            continue

        active_services = [
            cs for cs in customer.customer_services
            if cs.product_and_service.active and cs.rate and cs.rate > 0
        ]
        if not active_services:
            result.errors.append(
                f"Customer '{customer.display_name}' has no active services with valid rates — skipped."
            )
            continue

        center_id_to_invoice: dict[int, Invoice | None] = {
            center_by_name[n].id: None for n in customer_centers
        }
        for inv in invoices_by_company.get(company_id, []):
            for c in inv.centers:
                if c.id in center_id_to_invoice:
                    center_id_to_invoice[c.id] = inv

        group_map: dict[int | None, list[str]] = {}
        for name_lower in customer_centers:
            ctr = center_by_name[name_lower]
            inv = center_id_to_invoice.get(ctr.id)
            key = inv.id if inv else None
            group_map.setdefault(key, []).append(name_lower)

        for group_key, group_name_lowers in group_map.items():
            group_center_names = [center_by_name[n].name for n in group_name_lowers]

            if group_key is None:
                standalone_names = [center_by_name[n].name for n in group_name_lowers]
                _create_and_send(
                    db=db, qbo=qbo, access_token=access_token, realm_id=realm_id,
                    customer=customer, center_names=standalone_names,
                    center_by_name=center_by_name, parsed=parsed,
                    customer_services=active_services, inv_date=inv_date,
                    result=result, is_standalone=True,
                    invoice_upload_id=invoice_upload_id,
                    tax_code_id=tax_code_id,
                )
            else:
                _create_and_send(
                    db=db, qbo=qbo, access_token=access_token, realm_id=realm_id,
                    customer=customer, center_names=group_center_names,
                    center_by_name=center_by_name, parsed=parsed,
                    customer_services=active_services, inv_date=inv_date,
                    result=result, is_standalone=False,
                    invoice_upload_id=invoice_upload_id,
                    tax_code_id=tax_code_id,
                )

    return result


def generate_invoices(
    db: Session,
    qbo: SupportsQuickBooks,
    access_token: str,
    realm_id: str,
    filename: str,
    content: bytes,
    invoice_upload_id: int | None = None,
) -> GenerationResult:
    result = GenerationResult()

    # 1. Parse file
    parsed = parse_spreadsheet(filename, content)
    return generate_invoices_from_parsed(
        db=db, qbo=qbo, access_token=access_token, realm_id=realm_id,
        parsed=parsed, invoice_upload_id=invoice_upload_id,
    )


def _create_and_send(
    db: Session,
    qbo: SupportsQuickBooks,
    access_token: str,
    realm_id: str,
    customer: Customer,
    center_names: list[str],
    center_by_name: dict[str, Center],
    parsed: ParsedFile,
    customer_services: list[CustomerProductAndService],
    inv_date: date,
    result: GenerationResult,
    is_standalone: bool,
    invoice_upload_id: int | None = None,
    tax_code_id: str = "",
) -> None:
    if is_standalone and len(center_names) == 1:
        label = f"{center_names[0]} (standalone)"
    else:
        label = " + ".join(center_names)
    built = _build_qbo_invoice_payload(
        customer=customer,
        center_names=center_names,
        parsed=parsed,
        center_by_name=center_by_name,
        customer_services=customer_services,
        inv_date=inv_date,
        tax_code_id=tax_code_id,
    )
    if built is None:
        result.errors.append(
            f"Customer '{customer.display_name}' / {label}: no line items — invoice skipped."
        )
        return

    payload, total_amount, line_items = built

    try:
        qbo_inv = qbo.create_invoice(access_token, realm_id, payload)
        inv_id = str(qbo_inv.get("Id", ""))
        inv_number: str | None = qbo_inv.get("DocNumber") or None

        gen_inv_id: int | None = None
        if invoice_upload_id is not None and inv_id:
            gen_inv = GeneratedInvoice(
                invoice_upload_id=invoice_upload_id,
                customer_id=customer.id,
                center_group_name=label,
                total_amount=total_amount,
                send_status="pending",
                quickbooks_invoice_id=inv_id,
                invoice_number=inv_number,
            )
            db.add(gen_inv)
            db.flush()

            for name in center_names:
                ctr = center_by_name.get(name.lower())
                db.add(GeneratedInvoiceCenter(
                    generated_invoice_id=gen_inv.id,
                    center_id=ctr.id if ctr else None,
                    center_name=name,
                ))

            for li in line_items:
                db.add(GeneratedInvoiceLineItem(
                    generated_invoice_id=gen_inv.id,
                    product_and_service_id=li.product_and_service_id,
                    product_name=li.product_name,
                    description=li.description,
                    quantity=li.quantity,
                    rate=li.rate,
                    amount=li.amount,
                ))

            db.commit()
            gen_inv_id = gen_inv.id

        result.invoices_created += 1
        result.invoice_details.append(
            InvoiceDetail(
                customer_display_name=customer.display_name,
                group_description=label,
                qbo_invoice_id=inv_id,
                invoice_number=inv_number,
                sent_at=None,
                send_status="pending",
                generated_invoice_id=gen_inv_id,
            )
        )
    except Exception as err:
        result.invoices_failed += 1
        result.errors.append(
            f"Customer '{customer.display_name}' / {label}: failed to create invoice — {err}"
        )


# ── Dry-run line-item preview (for Sheet 2 download) ─────────────────────────

def build_line_item_preview(body: "RevalidateRequest", db: Session) -> dict:
    """
    Dry-run of invoice generation.  Returns the data needed to build the
    QBO-format Sheet 2 (invoice line items) without touching QBO or the DB.
    """
    from app.schemas.invoice_validation import RevalidateRequest  # local to avoid circular
    from app.services.invoice_validation import _rows_to_parsed_file

    parsed = _rows_to_parsed_file(body.rows, body.metric_columns)
    # Preview uses the PREVIOUS month's last day (invoices are for the period just ended)
    today = date.today()
    first_of_current = date(today.year, today.month, 1)
    inv_date = first_of_current - timedelta(days=1)  # last day of previous month
    due = _due_date(inv_date)
    memo = _memo(inv_date)
    month_label = _month_label(inv_date)

    # Last invoice number from DB
    last_inv = (
        db.query(GeneratedInvoice.invoice_number)
        .filter(GeneratedInvoice.invoice_number.isnot(None))
        .order_by(GeneratedInvoice.id.desc())
        .first()
    )
    last_invoice_no_str = last_inv[0] if last_inv else None
    try:
        last_invoice_no_int = int(last_invoice_no_str) if last_invoice_no_str else 0
    except ValueError:
        last_invoice_no_int = 0

    # Match centers to DB
    names_in_file = list(parsed.rows.keys())
    centers_in_db: list[Center] = (
        db.query(Center)
        .filter(sa_func.lower(Center.name).in_(names_in_file))
        .all()
    )
    center_by_name: dict[str, Center] = {c.name.lower(): c for c in centers_in_db}
    matched_names = [n for n in names_in_file if n in center_by_name]
    customer_ids = {center_by_name[n].company_id for n in matched_names}

    customers: list[Customer] = (
        db.query(Customer)
        .filter(Customer.id.in_(customer_ids))
        .options(
            selectinload(Customer.customer_services)
            .selectinload(CustomerProductAndService.product_and_service),
        )
        .all()
    )
    customer_by_id: dict[int, Customer] = {c.id: c for c in customers}

    invoices: list[Invoice] = (
        db.query(Invoice)
        .options(selectinload(Invoice.centers))
        .filter(Invoice.company_id.in_(customer_ids))
        .all()
    )
    invoices_by_company: dict[int, list[Invoice]] = {}
    for inv in invoices:
        invoices_by_company.setdefault(inv.company_id, []).append(inv)

    result_invoices: list[dict] = []
    inv_offset = 0

    for company_id in sorted(customer_ids):
        customer = customer_by_id.get(company_id)
        if not customer or not customer.qbo_id:
            continue

        customer_centers = [n for n in matched_names if center_by_name[n].company_id == company_id]
        if not customer_centers:
            continue

        active_services = [
            cs for cs in customer.customer_services
            if cs.product_and_service.active and cs.rate and cs.rate > 0
        ]
        if not active_services:
            continue

        center_id_to_inv: dict[int, Invoice | None] = {
            center_by_name[n].id: None for n in customer_centers
        }
        for inv in invoices_by_company.get(company_id, []):
            for c in inv.centers:
                if c.id in center_id_to_inv:
                    center_id_to_inv[c.id] = inv

        group_map: dict[int | None, list[str]] = {}
        for name_lower in customer_centers:
            ctr = center_by_name[name_lower]
            key = center_id_to_inv.get(ctr.id)
            group_map.setdefault(key.id if key else None, []).append(name_lower)

        for _group_key, group_names_lower in group_map.items():
            line_items: list[dict] = []
            for name_lower in group_names_lower:
                center_metrics = parsed.rows.get(name_lower, {})
                center_prefix = parsed.center_prefixes.get(name_lower, name_lower)
                items = _build_line_items_for_center(
                    center_prefix=center_prefix,
                    center_metrics=center_metrics,
                    customer_services=active_services,
                    month_label=month_label,
                )
                for li in items:
                    tax_amount = (li.amount * Decimal("0.1")).quantize(Decimal("0.0001"))
                    line_items.append({
                        "product_name": li.product_name,
                        "description": li.description,
                        "quantity": float(li.quantity),
                        "rate": float(li.rate),
                        "amount": float(li.amount),
                        "tax_amount": float(tax_amount),
                        "tax_code": "GST",
                    })

            if line_items:
                inv_offset += 1
                result_invoices.append({
                    "invoice_no": last_invoice_no_int + inv_offset,
                    "customer_display_name": customer.display_name,
                    "line_items": line_items,
                })

    return {
        "last_invoice_no": last_invoice_no_str,
        "invoice_date": inv_date.strftime("%d/%m/%Y"),
        "due_date": due.strftime("%d/%m/%Y"),
        "memo": memo,
        "invoices": result_invoices,
    }
